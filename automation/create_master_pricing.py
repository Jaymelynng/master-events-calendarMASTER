import json, os, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_fill = PatternFill('solid', fgColor='4A2C2A')
header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
active_fill = PatternFill('solid', fgColor='E8F5E9')
legacy_fill = PatternFill('solid', fgColor='FFF3E0')
unknown_fill = PatternFill('solid', fgColor='F5F5F5')
warn_fill = PatternFill('solid', fgColor='FFEBEE')
subheader_fill = PatternFill('solid', fgColor='8B6F6F')
subheader_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

GYM_MAP = {
    'capgymavery': ('CCP', 'Capital Gymnastics Cedar Park'),
    'capgymhp': ('CPF', 'Capital Gymnastics Pflugerville'),
    'capgymroundrock': ('CRR', 'Capital Gymnastics Round Rock'),
    'estrellagymnastics': ('EST', 'Estrella Gymnastics'),
    'houstongymnastics': ('HGA', 'Houston Gymnastics Academy'),
    'oasisgymnastics': ('OAS', 'Oasis Gymnastics'),
    'rbatascocita': ('RBA', 'Rowland Ballard Atascocita'),
    'rbkingwood': ('RBK', 'Rowland Ballard Kingwood'),
    'scottsdalegymnastics': ('SGT', 'Scottsdale Gymnastics'),
    'tigar': ('TIG', 'TIGAR Gymnastics'),
}

# Map schedule names to program types and active/legacy status
def classify_schedule(name, gym_code):
    n = name.upper().strip()

    # Birthday
    if 'BIRTHDAY' in n:
        return 'BIRTHDAY PARTY', None

    # Before/After Care
    if 'BEFORE CARE' in n or 'AFTER CARE' in n or n.startswith('AS PRE-REG'):
        return 'CAMP ADD-ON', None

    # Bring a Friend
    if 'BRING' in n and 'FRIEND' in n:
        return 'KNO (Bring-a-Friend)', None

    # KNO
    if 'KIDS NIGHT OUT' in n or "KID'S NIGHT OUT" in n or 'KID\u2019S NIGHT OUT' in n or 'KNO' in n or "PARENT'S NIGHT OUT" in n or 'PARENTS NIGHT OUT' in n or 'PARENTS DAY OUT' in n:
        if '2026' in n or '2025' in n:
            return 'KNO', 'ACTIVE-2026'
        if 'DAY OF' in n:
            return 'KNO', 'ACTIVE-DAY-OF'
        # Check price to determine if it's the new price
        return 'KNO', None

    # Open Gym
    if 'OPEN GYM' in n or 'GYM FUN FRIDAY' in n or 'FRIDAY FUN' in n or 'HOMESCHOOL OPEN GYM' in n:
        return 'OPEN GYM', None

    # Clinic
    if 'CLINIC' in n and 'CAMP' not in n:
        if '2026' in n:
            return 'CLINIC', 'ACTIVE-2026'
        return 'CLINIC', None

    # Showcase
    if 'SHOWCASE' in n:
        return 'SHOWCASE', None

    # Private lessons
    if 'PRIVATE' in n:
        return 'PRIVATE LESSON', None

    # Meet/Competition fees
    if 'MEET' in n or 'COMPULSORY' in n or 'IGT' in n or 'BOOTCAMP' in n:
        return 'COMPETITION/MEET', None

    # Summer camp
    if 'SUMMER CAMP' in n or 'SUMMER' in n and 'CAMP' in n:
        if '2026' in n:
            return 'SUMMER CAMP', 'ACTIVE-2026'
        if '2024' in n or '2023' in n:
            return 'SUMMER CAMP', 'LEGACY'
        return 'SUMMER CAMP', None

    # School year camp
    if 'SCHOOL YEAR' in n or 'SCHOOL' in n and 'CAMP' in n or "SCHOOL'S OUT" in n:
        if '2026' in n:
            return 'SCHOOL YEAR CAMP', 'ACTIVE-2026'
        return 'SCHOOL YEAR CAMP', None

    # Holiday/seasonal camps
    if 'SPRING BREAK' in n:
        if '2026' in n:
            return 'SPRING BREAK CAMP', 'ACTIVE-2026'
        return 'SPRING BREAK CAMP', None
    if 'CHRISTMAS' in n or 'WINTER' in n:
        return 'WINTER CAMP', None
    if 'THANKSGIVING' in n or 'HOLIDAY' in n:
        return 'HOLIDAY CAMP', None

    # General camp
    if 'CAMP' in n or 'FULL DAY' in n or 'HALF DAY' in n:
        if '2026' in n:
            return 'CAMP', 'ACTIVE-2026'
        if '2024' in n or '2023' in n:
            return 'CAMP', 'LEGACY'
        return 'CAMP', None

    # Special/Other
    if 'DANCE' in n:
        return 'DANCE', None
    if 'BOO BASH' in n or 'FALL FESTIVAL' in n or 'FLIP OR TREAT' in n:
        return 'SPECIAL EVENT', None
    if 'FREE' in n:
        return 'FREE EVENT', None
    if 'MOVIE' in n or 'PIZZA' in n:
        return 'SPECIAL EVENT', None
    if 'ADULT' in n or 'TUMBLING' in n and 'CLINIC' not in n:
        return 'OTHER', None
    if 'PRORATED' in n or 'PROMO' in n:
        return 'CAMP (PROMO)', None
    if 'MINI CAMP' in n or 'ZUNI' in n:
        return 'KNO VARIANT', None
    if 'SKILLS MEET' in n:
        return 'COMPETITION/MEET', None
    if 'OPEN HOUSE' in n:
        return 'SPECIAL EVENT', None
    if 'XCEL' in n or 'CALIBER' in n:
        return 'TEAM EVENT', None
    if 'SUPER CAMP' in n or 'JUNIOR CAMP' in n:
        return 'CAMP', None
    if '4TH OF JULY' in n:
        return 'CAMP', None

    return 'OTHER', None


# Kim's approved pricing (what it SHOULD be)
# CLINIC: $30/hr base but duration varies — can't flag as wrong without time info
KIMS_APPROVED = {
    'CLINIC': {'1st': 30, '2nd': 27, '3rd': 27, 'note': '$30/hr base — verify duration & if schedule is still active', 'no_flag': True},
    'KNO': {'1st': 45, '2nd': 40.50, '3rd': None, 'note': '$45, 10% sibling'},
}


def detect_sibling_discount(pricing_entry):
    p1 = pricing_entry.get('1st_child')
    p2 = pricing_entry.get('2nd_child')
    p3 = pricing_entry.get('3rd_child')

    if p1 is None or p1 == 0:
        return None, None, None, None, None

    disc2_type = disc2_amt = disc3_type = disc3_amt = None

    if p2 is not None:
        if p2 == p1:
            disc2_type = 'None'
            disc2_amt = 0
        elif p2 == 0:
            disc2_type = 'FREE'
            disc2_amt = p1
        else:
            diff = round(p1 - p2, 2)
            pct = round((1 - p2/p1) * 100, 1)
            if abs(pct - round(pct)) < 0.1 and pct in [5, 10, 15, 20, 25, 50]:
                disc2_type = f'{int(pct)}% off'
                disc2_amt = diff
            else:
                disc2_type = f'Flat ${diff:.0f} off' if diff == int(diff) else f'Flat ${diff:.2f} off'
                disc2_amt = diff

    if p3 is not None:
        if p3 == p1:
            disc3_type = 'None'
            disc3_amt = 0
        elif p3 == 0:
            disc3_type = 'FREE'
            disc3_amt = p1
        elif p2 is not None and p3 == p2:
            disc3_type = 'Same as 2nd'
            disc3_amt = round(p1 - p3, 2)
        else:
            diff = round(p1 - p3, 2)
            pct = round((1 - p3/p1) * 100, 1)
            if abs(pct - round(pct)) < 0.1 and pct in [5, 10, 15, 20, 25, 50]:
                disc3_type = f'{int(pct)}% off'
                disc3_amt = diff
            else:
                disc3_type = f'Flat ${diff:.0f} off' if diff == int(diff) else f'Flat ${diff:.2f} off'
                disc3_amt = diff

    return p2, p3, disc2_type, disc3_type, None


# ========== SHEET 1: MASTER LIST ==========
ws1 = wb.active
ws1.title = 'All Pricing Schedules'

headers = [
    'Gym', 'Gym Code', 'Schedule Name', 'Program Type', 'Status',
    '1st Child', '2nd Child', '3rd Child',
    '2nd Child Discount', '3rd Child Discount',
    'Per Day?', 'Weekly Price', 'Deposit',
    "Kim's Approved", 'Match?', 'Notes'
]

for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=i, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws1.row_dimensions[1].height = 30

download_dir = r'C:\Users\Jayme\Downloads'
files = glob.glob(os.path.join(download_dir, 'iclass-pricing-*.json'))

row = 2
all_data = []

for f in sorted(files):
    with open(f, 'r', encoding='utf-8') as fh:
        data = json.load(fh)

    slug = data['_meta']['gym_slug']
    if slug not in GYM_MAP:
        continue

    gym_code, gym_name = GYM_MAP[slug]

    for sched in data['pricing_schedules']:
        name = sched['name']
        pricing = sched['pricing']
        deposits = sched.get('deposits')

        program_type, status = classify_schedule(name, gym_code)

        # Get day-1 pricing
        day1 = next((p for p in pricing if p['days'] == 1), None)
        if not day1:
            day1 = pricing[0] if pricing else {}

        p1 = day1.get('1st_child')
        p2 = day1.get('2nd_child')
        p3 = day1.get('3rd_child')

        # Sibling discounts
        disc2_type = disc3_type = ''
        if p1 and p1 > 0 and p2 is not None:
            if p2 == p1:
                disc2_type = 'No discount'
            elif p2 == 0:
                disc2_type = 'FREE'
            else:
                diff = round(p1 - p2, 2)
                pct = round((1 - p2/p1) * 100, 1)
                if abs(pct - round(pct)) < 0.5 and round(pct) in [5, 10, 15, 20, 25, 50]:
                    disc2_type = f'{int(round(pct))}% off (=${diff:.2f})'
                else:
                    disc2_type = f'Flat ${diff:.2f} off ({pct:.1f}%)'

        if p1 and p1 > 0 and p3 is not None:
            if p3 == p1:
                disc3_type = 'No discount'
            elif p3 == 0:
                disc3_type = 'FREE'
            elif p2 is not None and p3 == p2:
                disc3_type = 'Same as 2nd'
            else:
                diff = round(p1 - p3, 2)
                pct = round((1 - p3/p1) * 100, 1)
                if abs(pct - round(pct)) < 0.5 and round(pct) in [5, 10, 15, 20, 25, 50]:
                    disc3_type = f'{int(round(pct))}% off (=${diff:.2f})'
                else:
                    disc3_type = f'Flat ${diff:.2f} off ({pct:.1f}%)'

        # Weekly price (5-day price if exists)
        day5 = next((p for p in pricing if p['days'] == 5), None)
        weekly = f"${day5['1st_child']:.0f}" if day5 and day5['1st_child'] != (p1 or 0) * 5 else ''
        if day5 and p1 and day5['1st_child'] < p1 * 5 and day5['1st_child'] != p1:
            weekly = f"${day5['1st_child']:.0f} (saves ${p1*5 - day5['1st_child']:.0f})"

        # Is it per-day pricing?
        is_per_day = len(pricing) > 1 and all(
            abs(p.get('1st_child', 0) - (p1 or 0) * p['days']) < 2
            for p in pricing[:3] if p['days'] <= 3
        )
        per_day_str = 'Yes' if is_per_day and len(pricing) > 1 else 'Flat rate' if len(pricing) == 1 else 'Tiered'

        # Deposit info
        dep_str = ''
        if deposits:
            d = deposits[0]
            if d['isPercent']:
                dep_str = f"{d['amount']}%"
            else:
                dep_str = f"${d['amount']:.0f}"

        # Status detection
        if status is None:
            if '2026' in name:
                status = 'ACTIVE-2026'
            elif '2024' in name or '2023' in name or '2022' in name or '2021' in name or '2019' in name:
                status = 'LEGACY'
            elif program_type in ['CLINIC', 'KNO', 'OPEN GYM', 'SHOWCASE', 'BIRTHDAY PARTY']:
                status = 'ACTIVE'
            elif 'PRORATED' in name.upper() or 'PROMO' in name.upper() or '10%' in name:
                status = 'PROMO/LEGACY'
            else:
                status = 'CHECK'

        # Kim's approved price comparison
        kim_str = ''
        match_str = ''
        approved = KIMS_APPROVED.get(program_type)
        if approved and 'ACTIVE' in status and 'Bring' not in name:
            kim_str = approved['note']
            if approved.get('no_flag'):
                # No duration info — can't determine if price is right or wrong
                if p1 is not None and p1 > 0:
                    multiple = round(p1 / approved['1st'], 1)
                    if multiple == 1.0:
                        match_str = f'${p1:.0f} (1hr?)'
                    elif multiple == int(multiple):
                        match_str = f'${p1:.0f} ({int(multiple)}hr?)'
                    else:
                        match_str = f'${p1:.0f} ({multiple}hr?)'
            elif p1 == approved['1st']:
                match_str = '✓'
            elif p1 is not None:
                match_str = f'WRONG — should be ${approved["1st"]:.0f}'
                if status == 'ACTIVE':
                    status = 'NEEDS UPDATE'

        notes = ''

        ws1.cell(row=row, column=1, value=gym_name)
        ws1.cell(row=row, column=2, value=gym_code)
        ws1.cell(row=row, column=3, value=name)
        ws1.cell(row=row, column=4, value=program_type)
        ws1.cell(row=row, column=5, value=status)
        ws1.cell(row=row, column=6, value=f'${p1:.2f}' if p1 is not None else '')
        ws1.cell(row=row, column=7, value=f'${p2:.2f}' if p2 is not None else '—')
        ws1.cell(row=row, column=8, value=f'${p3:.2f}' if p3 is not None else '—')
        ws1.cell(row=row, column=9, value=disc2_type)
        ws1.cell(row=row, column=10, value=disc3_type)
        ws1.cell(row=row, column=11, value=per_day_str)
        ws1.cell(row=row, column=12, value=weekly)
        ws1.cell(row=row, column=13, value=dep_str)
        ws1.cell(row=row, column=14, value=kim_str)
        ws1.cell(row=row, column=15, value=match_str)
        ws1.cell(row=row, column=16, value=notes)

        # Style
        if status == 'NEEDS UPDATE':
            fill = warn_fill
        elif 'ACTIVE' in status:
            fill = active_fill
        elif 'LEGACY' in status or 'PROMO' in status:
            fill = legacy_fill
        elif status == 'CHECK':
            fill = unknown_fill
        else:
            fill = None

        for c in range(1, 17):
            cell = ws1.cell(row=row, column=c)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if fill:
                cell.fill = fill
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row, column=9).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row, column=10).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row, column=14).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row, column=15).alignment = Alignment(horizontal='left', vertical='center')

        all_data.append({
            'gym_code': gym_code, 'gym_name': gym_name, 'name': name,
            'program': program_type, 'status': status,
            'p1': p1, 'p2': p2, 'p3': p3,
            'disc2': disc2_type, 'disc3': disc3_type,
            'weekly': weekly, 'deposit': dep_str,
            'kim': kim_str, 'match': match_str
        })

        row += 1

ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 8
ws1.column_dimensions['C'].width = 42
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 10
ws1.column_dimensions['H'].width = 10
ws1.column_dimensions['I'].width = 24
ws1.column_dimensions['J'].width = 24
ws1.column_dimensions['K'].width = 10
ws1.column_dimensions['L'].width = 22
ws1.column_dimensions['M'].width = 10
ws1.column_dimensions['N'].width = 22
ws1.column_dimensions['O'].width = 28
ws1.column_dimensions['P'].width = 30

# ========== SHEET 2: ACTIVE SUMMARY ==========
ws2 = wb.create_sheet('Active by Program')

ws2.merge_cells('A1:H1')
ws2['A1'] = 'ACTIVE PRICING SUMMARY — What Should Be Used Now'
ws2['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws2['A1'].alignment = Alignment(horizontal='center')

programs = ['CLINIC', 'KNO', 'OPEN GYM', 'CAMP', 'SUMMER CAMP', 'SCHOOL YEAR CAMP', 'SPRING BREAK CAMP', 'HOLIDAY CAMP', 'WINTER CAMP', 'SHOWCASE', 'BIRTHDAY PARTY']
gym_order = ['CCP', 'CPF', 'CRR', 'EST', 'HGA', 'OAS', 'RBA', 'RBK', 'SGT', 'TIG']

r = 3
for prog in programs:
    ws2.cell(row=r, column=1, value=prog)
    ws2.merge_cells(f'A{r}:H{r}')
    for c in range(1, 9):
        cell = ws2.cell(row=r, column=c)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = thin_border
    r += 1

    sub_headers = ['Gym', 'Schedule Name', '1st Child', '2nd Child', '3rd Child', 'Sibling Discount', 'Weekly', 'Status']
    for i, h in enumerate(sub_headers, 1):
        cell = ws2.cell(row=r, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    r += 1

    found = False
    for gc in gym_order:
        matches = [d for d in all_data if d['gym_code'] == gc and d['program'] == prog]
        # Prefer ACTIVE-2026, then ACTIVE
        active = [d for d in matches if 'ACTIVE' in (d['status'] or '')]
        if not active:
            active = [d for d in matches if d['status'] not in ['LEGACY', 'PROMO/LEGACY']]

        for d in active:
            found = True
            ws2.cell(row=r, column=1, value=f"{d['gym_name']} ({d['gym_code']})")
            ws2.cell(row=r, column=2, value=d['name'])
            ws2.cell(row=r, column=3, value=f"${d['p1']:.2f}" if d['p1'] is not None else '')
            ws2.cell(row=r, column=4, value=f"${d['p2']:.2f}" if d['p2'] is not None else '—')
            ws2.cell(row=r, column=5, value=f"${d['p3']:.2f}" if d['p3'] is not None else '—')
            ws2.cell(row=r, column=6, value=d['disc2'] if d['disc2'] else '—')
            ws2.cell(row=r, column=7, value=d['weekly'])
            ws2.cell(row=r, column=8, value=d['status'])

            fill = active_fill if 'ACTIVE' in (d['status'] or '') else unknown_fill
            for c in range(1, 9):
                cell = ws2.cell(row=r, column=c)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = fill
            ws2.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws2.cell(row=r, column=6).alignment = Alignment(horizontal='left', vertical='center')
            r += 1

    if not found:
        ws2.cell(row=r, column=1, value='No active schedules found')
        ws2.cell(row=r, column=1).font = Font(name='Arial', size=10, italic=True, color='999999')
        r += 1

    r += 1

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 42
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 26
ws2.column_dimensions['G'].width = 22
ws2.column_dimensions['H'].width = 14

# ========== SHEET 3: SIBLING DISCOUNT SUMMARY ==========
ws3 = wb.create_sheet('Sibling Discounts')

ws3.merge_cells('A1:F1')
ws3['A1'] = 'SIBLING DISCOUNT PATTERNS BY GYM'
ws3['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws3['A1'].alignment = Alignment(horizontal='center')

r = 3
sib_headers = ['Gym', 'Program', 'Schedule', '2nd Child Discount', '3rd Child Discount', 'Type']
for i, h in enumerate(sib_headers, 1):
    cell = ws3.cell(row=r, column=i, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
r += 1

for gc in gym_order:
    relevant = [d for d in all_data if d['gym_code'] == gc and d['disc2'] and d['disc2'] != 'No discount' and 'ACTIVE' in (d['status'] or '') or (d['gym_code'] == gc and d['program'] in ['KNO', 'CLINIC', 'OPEN GYM', 'CAMP', 'SUMMER CAMP'] and d['disc2'] and d['disc2'] not in ['No discount', ''])]

    if not relevant:
        relevant = [d for d in all_data if d['gym_code'] == gc and d['disc2'] and d['disc2'] not in ['No discount', '']]

    for d in relevant:
        ws3.cell(row=r, column=1, value=f"{d['gym_name']} ({d['gym_code']})")
        ws3.cell(row=r, column=2, value=d['program'])
        ws3.cell(row=r, column=3, value=d['name'])
        ws3.cell(row=r, column=4, value=d['disc2'])
        ws3.cell(row=r, column=5, value=d['disc3'] if d['disc3'] else '—')

        if '%' in (d['disc2'] or ''):
            dtype = 'Percentage'
        elif 'Flat' in (d['disc2'] or ''):
            dtype = 'Flat Rate'
        elif 'FREE' in (d['disc2'] or ''):
            dtype = 'Free'
        else:
            dtype = '—'
        ws3.cell(row=r, column=6, value=dtype)

        for c in range(1, 7):
            cell = ws3.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws3.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws3.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws3.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='center')
        r += 1

ws3.column_dimensions['A'].width = 38
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 42
ws3.column_dimensions['D'].width = 26
ws3.column_dimensions['E'].width = 26
ws3.column_dimensions['F'].width = 14

# ========== SHEET 4: LEGACY/CLEANUP ==========
ws4 = wb.create_sheet('Legacy - Needs Cleanup')

ws4.merge_cells('A1:E1')
ws4['A1'] = 'LEGACY PRICING SCHEDULES — Should These Be Deleted?'
ws4['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws4['A1'].alignment = Alignment(horizontal='center')

r = 3
leg_headers = ['Gym', 'Schedule Name', 'Program Type', 'Price', 'Why Legacy?']
for i, h in enumerate(leg_headers, 1):
    cell = ws4.cell(row=r, column=i, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
r += 1

for d in all_data:
    if d['status'] in ['LEGACY', 'PROMO/LEGACY']:
        ws4.cell(row=r, column=1, value=f"{d['gym_name']} ({d['gym_code']})")
        ws4.cell(row=r, column=2, value=d['name'])
        ws4.cell(row=r, column=3, value=d['program'])
        ws4.cell(row=r, column=4, value=f"${d['p1']:.2f}" if d['p1'] is not None else '')

        reason = ''
        if '2024' in d['name'] or '2023' in d['name'] or '2022' in d['name']:
            reason = f"Old year in name"
        elif '10%' in d['name']:
            reason = 'Promo/discount schedule'
        elif 'PRORATED' in d['name'].upper() or 'PROMO' in d['name'].upper():
            reason = 'Promo pricing'
        ws4.cell(row=r, column=5, value=reason)

        for c in range(1, 6):
            cell = ws4.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = legacy_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws4.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws4.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
        r += 1

ws4.column_dimensions['A'].width = 38
ws4.column_dimensions['B'].width = 45
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 30

# Legend on first sheet
r = ws1.max_row + 2
ws1.cell(row=r, column=1, value='Legend:').font = bold_font
r += 1
ws1.cell(row=r, column=1, value='Green = Active / Current').font = normal_font
ws1.cell(row=r, column=2).fill = active_fill
r += 1
ws1.cell(row=r, column=1, value='Orange = Legacy / Old Year').font = normal_font
ws1.cell(row=r, column=2).fill = legacy_fill
r += 1
ws1.cell(row=r, column=1, value='Gray = Needs Review').font = normal_font
ws1.cell(row=r, column=2).fill = unknown_fill

out = r'C:\Users\Jayme\Downloads\iClass_Master_Pricing_Data_Collection.xlsx'
wb.save(out)
print(f'Saved to {out}')
print(f'Total schedules: {len(all_data)}')
for gc in gym_order:
    count = len([d for d in all_data if d['gym_code'] == gc])
    active = len([d for d in all_data if d['gym_code'] == gc and 'ACTIVE' in (d['status'] or '')])
    legacy = len([d for d in all_data if d['gym_code'] == gc and d['status'] in ['LEGACY', 'PROMO/LEGACY']])
    print(f'  {gc}: {count} total, {active} active, {legacy} legacy')
