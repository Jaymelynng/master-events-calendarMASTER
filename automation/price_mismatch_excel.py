#!/usr/bin/env python3
"""
Create Excel file with clickable hyperlinks for price mismatches.
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Source of Truth - CONFIRMED CORRECT PRICES
CORRECT_PRICES = {
    "Capital Gymnastics - Cedar Park": {
        "CLINIC": {
            "default": 25,
            "bar": 30,
            "dance_1.5": 35,
            "flip_flop": 20,
            "cartwheel_20": 20,
            "hot_shots": 50,
        },
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": {
            "gym_fun_fridays": 10,
            "open_gym": 20,
        }
    },
    "Capital Gymnastics - Pflugerville": {
        "CLINIC": {
            "default": 25,
            "bar": 30,
            "dance_1.5": 35,
            "flip_flop": 20,
        },
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": {
            "gym_fun_fridays": 10,
            "open_gym": 20,
        }
    },
    "Capital Gymnastics - Round Rock": {
        "CLINIC": {
            "1hr": 25,
            "1.5hr": 30,
            "2hr": 40,
            "specialty_1hr": 30,
            "specialty_1.5hr": 35,
            "specialty_2hr": 45,
        },
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": 10,
    },
    "Estrella Gymnastics": {
        "CLINIC": 25,
        "KIDS NIGHT OUT": 40,
        "OPEN GYM": {
            "homeschool": 10,
            "regular": 30,
        }
    },
    "Houston Gymnastics Academy": {
        "CLINIC": {
            "default": 25,
            "tumble_dance": 75,
            "conditioning": 30,
        },
        "KIDS NIGHT OUT": 40,
        "OPEN GYM": 20,
    },
    "Oasis Gymnastics": {
        "CLINIC": 25,
        "KIDS NIGHT OUT": 40,
        "OPEN GYM": 20,
    },
    "Rowland Ballard - Atascocita": {
        "CLINIC": {
            "default": 25,
            "team": 15,
        },
        "KIDS NIGHT OUT": [20, 25, 30, 35, 40],
        "OPEN GYM": 20,
    },
    "Rowland Ballard - Kingwood": {
        "CLINIC": {
            "back_handspring": 25,
            "bonus_tumbling": 15,
            "master_class": 30,
        },
        "KIDS NIGHT OUT": [35, 40, 45],
        "OPEN GYM": {
            "friday_fun_night": 15,
            "friday_fun_gym": [20, 25],
        }
    },
    "Scottsdale Gymnastics": {
        "CLINIC": {
            "sgt": 25,
            "spf": [40, 50],
            "tumbling": 40,
            "trampoline": 40,
        },
        "KIDS NIGHT OUT": 45,
        "OPEN GYM": 20,
    },
    "TIGAR Gymnastics": {
        "CLINIC": 25,
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": 20,
    }
}

GYM_NAMES = {
    'CCP': 'Capital Gymnastics - Cedar Park',
    'CPF': 'Capital Gymnastics - Pflugerville', 
    'CRR': 'Capital Gymnastics - Round Rock',
    'EST': 'Estrella Gymnastics',
    'HGA': 'Houston Gymnastics Academy',
    'OAS': 'Oasis Gymnastics',
    'RBA': 'Rowland Ballard - Atascocita',
    'RBK': 'Rowland Ballard - Kingwood',
    'SGT': 'Scottsdale Gymnastics',
    'TIG': 'TIGAR Gymnastics'
}

def get_correct_price(gym_name, event_type, title):
    """Get the CORRECT price from Source of Truth."""
    title_lower = title.lower()
    
    gym_prices = CORRECT_PRICES.get(gym_name, {})
    type_prices = gym_prices.get(event_type, None)
    
    if type_prices is None:
        return None
    
    if isinstance(type_prices, (int, float)):
        return type_prices
    
    if isinstance(type_prices, list):
        return type_prices[0]
    
    if isinstance(type_prices, dict):
        if 'bar' in title_lower:
            return type_prices.get('bar', type_prices.get('default'))
        if 'dance' in title_lower and ('1.5' in title_lower or '1 1/2' in title_lower):
            return type_prices.get('dance_1.5', type_prices.get('default'))
        if 'flip' in title_lower and 'flop' in title_lower:
            return type_prices.get('flip_flop', type_prices.get('default'))
        if 'hot shot' in title_lower:
            return type_prices.get('hot_shots', type_prices.get('default'))
        if 'tumble' in title_lower and 'dance' in title_lower:
            return type_prices.get('tumble_dance', type_prices.get('default'))
        if 'conditioning' in title_lower:
            return type_prices.get('conditioning', type_prices.get('default'))
        if 'master class' in title_lower:
            return type_prices.get('master_class', type_prices.get('default'))
        if 'gym fun' in title_lower or 'gff' in title_lower:
            return type_prices.get('gym_fun_fridays', type_prices.get('default'))
        if 'homeschool' in title_lower:
            return type_prices.get('homeschool', type_prices.get('default'))
        if 'friday fun' in title_lower:
            val = type_prices.get('friday_fun_gym', type_prices.get('friday_fun_night'))
            if isinstance(val, list):
                return val[0]
            return val
        
        if 'default' in type_prices:
            return type_prices['default']
        if '1hr' in type_prices:
            return type_prices['1hr']
        if 'open_gym' in type_prices:
            return type_prices['open_gym']
        if 'regular' in type_prices:
            return type_prices['regular']
        
        return list(type_prices.values())[0]
    
    return None

def main():
    json_path = r'c:\Users\Jayme\Downloads\events_rows - 2026-01-05T164936.423.json'
    output_xlsx = r'C:\JAYME PROJECTS\HUB -ACTIVE - master-events-calendar\exports\PRICE_MISMATCHES_TO_FIX.xlsx'
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    special_events = [e for e in data if e.get('type') in ['CLINIC', 'KIDS NIGHT OUT', 'OPEN GYM']]
    
    mismatches = []
    
    for e in special_events:
        gym_code = e.get('gym_id', 'unknown')
        gym_name = GYM_NAMES.get(gym_code, gym_code)
        title = e.get('title', '')
        date = e.get('date', '')
        etype = e.get('type', '')
        current_price = e.get('price')
        url = e.get('event_url', '')
        
        correct_price = get_correct_price(gym_name, etype, title)
        
        if correct_price is None:
            continue
        
        current_float = float(current_price) if current_price else None
        correct_float = float(correct_price)
        
        if current_float == correct_float:
            continue
        
        if current_float is None:
            continue
        
        clean_title = title.encode('ascii', 'ignore').decode('ascii')
        
        mismatches.append({
            'gym': gym_name,
            'type': etype,
            'date': date,
            'title': clean_title,
            'wrong_price': f"${int(current_float)}",
            'correct_price': f"${int(correct_float)}",
            'url': url
        })
    
    mismatches.sort(key=lambda x: (x['gym'], x['date']))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Mismatches"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    correct_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Gym', 'Event Type', 'Date', 'Event Title', 'WRONG Price (in iClass)', 'CORRECT Price', 'Click to Fix']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, m in enumerate(mismatches, 2):
        ws.cell(row=row_idx, column=1, value=m['gym']).border = thin_border
        ws.cell(row=row_idx, column=2, value=m['type']).border = thin_border
        ws.cell(row=row_idx, column=3, value=m['date']).border = thin_border
        ws.cell(row=row_idx, column=4, value=m['title'][:60]).border = thin_border
        
        wrong_cell = ws.cell(row=row_idx, column=5, value=m['wrong_price'])
        wrong_cell.fill = wrong_fill
        wrong_cell.border = thin_border
        wrong_cell.alignment = Alignment(horizontal='center')
        
        correct_cell = ws.cell(row=row_idx, column=6, value=m['correct_price'])
        correct_cell.fill = correct_fill
        correct_cell.border = thin_border
        correct_cell.alignment = Alignment(horizontal='center')
        
        # Clickable hyperlink
        link_cell = ws.cell(row=row_idx, column=7, value="CLICK TO FIX")
        link_cell.hyperlink = m['url']
        link_cell.font = link_font
        link_cell.border = thin_border
        link_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    col_widths = [35, 15, 12, 50, 20, 15, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_xlsx)
    print(f"Excel file saved to: {output_xlsx}")
    print(f"Total mismatches: {len(mismatches)}")
    print("\nOpen in Excel - the 'Click to Fix' links are clickable!")

if __name__ == "__main__":
    main()

