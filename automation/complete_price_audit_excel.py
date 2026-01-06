#!/usr/bin/env python3
"""
COMPLETE Price Audit Excel - ALL issues in one file:
1. Price mismatches (wrong price vs Source of Truth)
2. No description at all
3. Price in title but NOT in description
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Source of Truth - CONFIRMED CORRECT PRICES
CORRECT_PRICES = {
    "Capital Gymnastics - Cedar Park": {
        "CLINIC": {"default": 25, "bar": 30, "dance_1.5": 35, "flip_flop": 20, "hot_shots": 50},
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": {"gym_fun_fridays": 10, "open_gym": 20}
    },
    "Capital Gymnastics - Pflugerville": {
        "CLINIC": {"default": 25, "bar": 30, "dance_1.5": 35, "flip_flop": 20},
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": {"gym_fun_fridays": 10, "open_gym": 20}
    },
    "Capital Gymnastics - Round Rock": {
        "CLINIC": {"1hr": 25, "1.5hr": 30, "2hr": 40},
        "KIDS NIGHT OUT": 35,
        "OPEN GYM": 10
    },
    "Estrella Gymnastics": {
        "CLINIC": 25, "KIDS NIGHT OUT": 40, "OPEN GYM": {"homeschool": 10, "regular": 30}
    },
    "Houston Gymnastics Academy": {
        "CLINIC": {"default": 25, "tumble_dance": 75, "conditioning": 30},
        "KIDS NIGHT OUT": 40, "OPEN GYM": 20
    },
    "Oasis Gymnastics": {"CLINIC": 25, "KIDS NIGHT OUT": 40, "OPEN GYM": 20},
    "Rowland Ballard - Atascocita": {
        "CLINIC": {"default": 25, "team": 15},
        "KIDS NIGHT OUT": [20, 25, 30, 35, 40], "OPEN GYM": 20
    },
    "Rowland Ballard - Kingwood": {
        "CLINIC": {"back_handspring": 25, "bonus_tumbling": 15, "master_class": 30},
        "KIDS NIGHT OUT": [35, 40, 45],
        "OPEN GYM": {"friday_fun_night": 15, "friday_fun_gym": [20, 25]}
    },
    "Scottsdale Gymnastics": {
        "CLINIC": {"sgt": 25, "spf": [40, 50], "tumbling": 40, "trampoline": 40},
        "KIDS NIGHT OUT": 45, "OPEN GYM": 20
    },
    "TIGAR Gymnastics": {"CLINIC": 25, "KIDS NIGHT OUT": 35, "OPEN GYM": 20}
}

GYM_NAMES = {
    'CCP': 'Capital Gymnastics - Cedar Park',
    'CPF': 'Capital Gymnastics - Pflugerville', 
    'CRR': 'Capital Gymnastics - Round Rock',
    'EST': 'Estrella Gymnastics',
    'HGA': 'Houston Gymnastics Academy',
    'OAS': 'Oasis Gymnastics',
    'RBA': 'Rowland Ballard - Atascocita',
    'RBK': 'Rowland Ballard - Kingwood',
    'SGT': 'Scottsdale Gymnastics',
    'TIG': 'TIGAR Gymnastics'
}

def extract_prices(text):
    """Extract dollar amounts from text."""
    if not text:
        return []
    prices = re.findall(r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', text)
    return [float(p.replace(',', '')) for p in prices]

def get_correct_price(gym_name, event_type, title):
    """Get the CORRECT price from Source of Truth."""
    title_lower = title.lower()
    gym_prices = CORRECT_PRICES.get(gym_name, {})
    type_prices = gym_prices.get(event_type, None)
    
    if type_prices is None:
        return None
    if isinstance(type_prices, (int, float)):
        return type_prices
    if isinstance(type_prices, list):
        return type_prices[0]
    if isinstance(type_prices, dict):
        if 'bar' in title_lower:
            return type_prices.get('bar', type_prices.get('default'))
        if 'flip' in title_lower and 'flop' in title_lower:
            return type_prices.get('flip_flop', type_prices.get('default'))
        if 'gym fun' in title_lower or 'gff' in title_lower:
            return type_prices.get('gym_fun_fridays', type_prices.get('default'))
        if 'homeschool' in title_lower:
            return type_prices.get('homeschool', type_prices.get('default'))
        if 'friday fun' in title_lower:
            val = type_prices.get('friday_fun_gym', type_prices.get('friday_fun_night'))
            return val[0] if isinstance(val, list) else val
        if 'default' in type_prices:
            return type_prices['default']
        if '1hr' in type_prices:
            return type_prices['1hr']
        if 'open_gym' in type_prices:
            return type_prices['open_gym']
        if 'regular' in type_prices:
            return type_prices['regular']
        return list(type_prices.values())[0]
    return None

def main():
    json_path = r'c:\Users\Jayme\Downloads\events_rows - 2026-01-05T164936.423.json'
    output_xlsx = r'C:\JAYME PROJECTS\HUB -ACTIVE - master-events-calendar\exports\COMPLETE_PRICE_AUDIT.xlsx'
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    special_events = [e for e in data if e.get('type') in ['CAMP', 'CLINIC', 'KIDS NIGHT OUT', 'OPEN GYM']]
    
    # Categorize all issues
    wrong_price = []      # Price doesn't match SOT
    no_description = []   # No description at all
    price_not_in_desc = [] # Price in title but NOT in description
    
    for e in special_events:
        gym_code = e.get('gym_id', 'unknown')
        gym_name = GYM_NAMES.get(gym_code, gym_code)
        title = e.get('title', '')
        date = e.get('date', '')
        etype = e.get('type', '')
        current_price = e.get('price')
        desc = e.get('description', '') or ''
        url = e.get('event_url', '')
        
        clean_title = title.encode('ascii', 'ignore').decode('ascii')
        
        # Check for no description
        if not desc or len(desc.strip()) < 10:
            no_description.append({
                'gym': gym_name,
                'type': etype,
                'date': date,
                'title': clean_title,
                'price_in_title': f"${int(current_price)}" if current_price else "None",
                'url': url,
                'issue': 'NO DESCRIPTION'
            })
            continue
        
        # Check if price is in description
        title_prices = extract_prices(title)
        desc_prices = extract_prices(desc)
        
        if title_prices and not desc_prices:
            # Price in title but NOT in description
            price_not_in_desc.append({
                'gym': gym_name,
                'type': etype,
                'date': date,
                'title': clean_title,
                'price_in_title': f"${int(title_prices[0])}",
                'price_in_desc': "MISSING",
                'url': url,
                'issue': 'PRICE NOT IN DESCRIPTION'
            })
        elif not desc_prices:
            # No price anywhere in description
            price_not_in_desc.append({
                'gym': gym_name,
                'type': etype,
                'date': date,
                'title': clean_title,
                'price_in_title': f"${int(title_prices[0])}" if title_prices else "None",
                'price_in_desc': "MISSING",
                'url': url,
                'issue': 'PRICE NOT IN DESCRIPTION'
            })
        
        # Check for price mismatch (only for non-CAMP events)
        if etype != 'CAMP':
            correct_price = get_correct_price(gym_name, etype, title)
            if correct_price and current_price:
                current_float = float(current_price)
                correct_float = float(correct_price)
                if current_float != correct_float:
                    wrong_price.append({
                        'gym': gym_name,
                        'type': etype,
                        'date': date,
                        'title': clean_title,
                        'wrong_price': f"${int(current_float)}",
                        'correct_price': f"${int(correct_float)}",
                        'url': url,
                        'issue': 'WRONG PRICE'
                    })
    
    # Create Excel workbook with multiple sheets
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_fill_yellow = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    correct_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    missing_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: Wrong Price ==========
    ws1 = wb.active
    ws1.title = "1-WRONG PRICE"
    
    headers1 = ['Gym', 'Type', 'Date', 'Event Title', 'WRONG Price', 'CORRECT Price', 'Click to Fix']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill_red
        cell.border = thin_border
    
    for row_idx, m in enumerate(sorted(wrong_price, key=lambda x: (x['gym'], x['date'])), 2):
        ws1.cell(row=row_idx, column=1, value=m['gym']).border = thin_border
        ws1.cell(row=row_idx, column=2, value=m['type']).border = thin_border
        ws1.cell(row=row_idx, column=3, value=m['date']).border = thin_border
        ws1.cell(row=row_idx, column=4, value=m['title'][:50]).border = thin_border
        
        wrong_cell = ws1.cell(row=row_idx, column=5, value=m['wrong_price'])
        wrong_cell.fill = wrong_fill
        wrong_cell.border = thin_border
        
        correct_cell = ws1.cell(row=row_idx, column=6, value=m['correct_price'])
        correct_cell.fill = correct_fill
        correct_cell.border = thin_border
        
        link_cell = ws1.cell(row=row_idx, column=7, value="FIX IT")
        link_cell.hyperlink = m['url']
        link_cell.font = link_font
        link_cell.border = thin_border
    
    for col, w in enumerate([30, 12, 12, 45, 12, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'
    
    # ========== SHEET 2: No Description ==========
    ws2 = wb.create_sheet("2-NO DESCRIPTION")
    
    headers2 = ['Gym', 'Type', 'Date', 'Event Title', 'Price in Title', 'Click to Fix']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill_orange
        cell.border = thin_border
    
    for row_idx, m in enumerate(sorted(no_description, key=lambda x: (x['gym'], x['date'])), 2):
        ws2.cell(row=row_idx, column=1, value=m['gym']).border = thin_border
        ws2.cell(row=row_idx, column=2, value=m['type']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=m['date']).border = thin_border
        ws2.cell(row=row_idx, column=4, value=m['title'][:50]).border = thin_border
        ws2.cell(row=row_idx, column=5, value=m['price_in_title']).border = thin_border
        
        link_cell = ws2.cell(row=row_idx, column=6, value="ADD DESC")
        link_cell.hyperlink = m['url']
        link_cell.font = link_font
        link_cell.border = thin_border
    
    for col, w in enumerate([30, 12, 12, 45, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'
    
    # ========== SHEET 3: Price Not in Description ==========
    ws3 = wb.create_sheet("3-PRICE NOT IN DESC")
    
    headers3 = ['Gym', 'Type', 'Date', 'Event Title', 'Price in Title', 'Price in Desc', 'Click to Fix']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill_yellow
        cell.border = thin_border
    
    for row_idx, m in enumerate(sorted(price_not_in_desc, key=lambda x: (x['gym'], x['date'])), 2):
        ws3.cell(row=row_idx, column=1, value=m['gym']).border = thin_border
        ws3.cell(row=row_idx, column=2, value=m['type']).border = thin_border
        ws3.cell(row=row_idx, column=3, value=m['date']).border = thin_border
        ws3.cell(row=row_idx, column=4, value=m['title'][:50]).border = thin_border
        ws3.cell(row=row_idx, column=5, value=m['price_in_title']).border = thin_border
        
        missing_cell = ws3.cell(row=row_idx, column=6, value=m['price_in_desc'])
        missing_cell.fill = missing_fill
        missing_cell.border = thin_border
        
        link_cell = ws3.cell(row=row_idx, column=7, value="ADD PRICE")
        link_cell.hyperlink = m['url']
        link_cell.font = link_font
        link_cell.border = thin_border
    
    for col, w in enumerate([30, 12, 12, 45, 12, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = 'A2'
    
    wb.save(output_xlsx)
    
    print(f"Excel file saved to: {output_xlsx}")
    print()
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Sheet 1 - WRONG PRICE: {len(wrong_price)} events")
    print(f"Sheet 2 - NO DESCRIPTION: {len(no_description)} events")
    print(f"Sheet 3 - PRICE NOT IN DESCRIPTION: {len(price_not_in_desc)} events")
    print(f"\nTOTAL ISSUES: {len(wrong_price) + len(no_description) + len(price_not_in_desc)}")

if __name__ == "__main__":
    main()

