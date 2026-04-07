from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_fill = PatternFill('solid', fgColor='4A2C2A')
header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
subheader_fill = PatternFill('solid', fgColor='8B6F6F')
subheader_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
confirmed_fill = PatternFill('solid', fgColor='E8F5E9')
waiting_fill = PatternFill('solid', fgColor='FFF8E1')
no_change_fill = PatternFill('solid', fgColor='F5F5F5')
increase_font = Font(color='2E7D32', bold=True, name='Arial', size=10)
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_subheader(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill

# ========== SHEET 1: PRICING CHANGES ==========
ws1 = wb.active
ws1.title = 'Pricing Changes'

ws1.merge_cells('A1:G1')
ws1['A1'] = 'RISE ATHLETICS - Alternate Program Pricing Changes'
ws1['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:G2')
ws1['A2'] = 'Price Increases Effective March 2026 | Birthday Parties & Open Gym NOT included'
ws1['A2'].font = Font(size=10, name='Arial', color='888888', italic=True)
ws1['A2'].alignment = Alignment(horizontal='center')

# CLINIC section
r = 4
ws1.cell(row=r, column=1, value='CLINIC PRICING')
style_subheader(ws1, r, 7)
ws1.merge_cells(f'A{r}:G{r}')

r = 5
headers = ['Gym', 'Gym Code', 'Old Price', 'New Price', 'Change', 'Notes', 'Status']
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, 7)

# OLD = detailed list, NEW = Kim's table
clinic_data = [
    ['Capital Gymnastics Cedar Park', 'CCP', '$30 (1hr) / $35 (1.5hr)', '$35', '+$5 (1hr)', 'Was tiered, now flat $35', 'Confirmed'],
    ['Capital Gymnastics Pflugerville', 'CPF', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Capital Gymnastics Round Rock', 'CRR', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Estrella Gymnastics', 'EST', '---', '$30', '---', 'No old clinic listed', 'Confirmed'],
    ['Houston Gymnastics Academy', 'HGA', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Oasis Gymnastics', 'OAS', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Rowland Ballard Atascocita', 'RBA', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Rowland Ballard Kingwood', 'RBK', '$25', '$30', '+$5', '', 'Confirmed'],
    ['Scottsdale Gymnastics', 'SGT', '$25', '$30', '+$5', '', 'Confirmed'],
    ['TIGAR Gymnastics', 'TIG', '$25', '$30', '+$5', '', 'Confirmed'],
]

for i, row_data in enumerate(clinic_data):
    r = 6 + i
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    fill = confirmed_fill
    style_row(ws1, r, 7, fill)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
    if row_data[4] and row_data[4].startswith('+'):
        ws1.cell(row=r, column=5).font = increase_font

# KNO section
r = 17
ws1.cell(row=r, column=1, value='KIDS NIGHT OUT PRICING')
style_subheader(ws1, r, 7)
ws1.merge_cells(f'A{r}:G{r}')

r = 18
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, 7)

kno_data = [
    ['Capital Gymnastics Cedar Park', 'CCP', '$35', '$40', '+$5', 'Sibling: $30/sibling', 'Confirmed'],
    ['Capital Gymnastics Pflugerville', 'CPF', '$35', '$40', '+$5', 'Sibling: $30 for 2nd child', 'Confirmed'],
    ['Capital Gymnastics Round Rock', 'CRR', '$35', '$40', '+$5', 'Sibling: $30/sibling', 'Confirmed'],
    ['Estrella Gymnastics', 'EST', '$40', '$40', 'No change', '', 'Confirmed'],
    ['Houston Gymnastics Academy', 'HGA', '$40', '$45', '+$5', 'Sibling: $35 each for 2+', 'Confirmed'],
    ['Oasis Gymnastics', 'OAS', '$40', '$45', '+$5', '', 'Confirmed'],
    ['Rowland Ballard Atascocita', 'RBA', '$35', '$40', '+$5', 'No sibling discounts', 'Confirmed'],
    ['Rowland Ballard Kingwood', 'RBK', '$35 (std) / $40 (ext)', '$40', '+$5', '', 'Confirmed'],
    ['Scottsdale Gymnastics', 'SGT', '$45', '$45', 'No change', '', 'Confirmed'],
    ['TIGAR Gymnastics', 'TIG', '$35', '$40', '+$5', 'No sibling discounts', 'Confirmed'],
]

for i, row_data in enumerate(kno_data):
    r = 19 + i
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    if row_data[4] == 'No change':
        fill = no_change_fill
    else:
        fill = confirmed_fill
    style_row(ws1, r, 7, fill)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
    if row_data[4] and row_data[4].startswith('+'):
        ws1.cell(row=r, column=5).font = increase_font

# Legend
r = 30
ws1.cell(row=r, column=1, value='Legend:').font = bold_font
r += 1
ws1.cell(row=r, column=1, value='Green = Price increase confirmed').font = normal_font
ws1.cell(row=r, column=2).fill = confirmed_fill
r += 1
ws1.cell(row=r, column=1, value='Gray = No Change').font = normal_font
ws1.cell(row=r, column=2).fill = no_change_fill

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 24
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 30
ws1.column_dimensions['G'].width = 14

# ========== SHEET 2: OPEN GYM ==========
ws2 = wb.create_sheet('Current Open Gym')

ws2.merge_cells('A1:D1')
ws2['A1'] = 'OPEN GYM PRICING - No Changes'
ws2['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws2['A1'].alignment = Alignment(horizontal='center')

r = 3
for i, h in enumerate(['Gym', 'Gym Code', 'Price', 'Status'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 4)

og_data = [
    ['Capital Gymnastics Cedar Park', 'CCP', '$10'],
    ['Capital Gymnastics Pflugerville', 'CPF', '$10'],
    ['Capital Gymnastics Round Rock', 'CRR', '$10'],
    ['Estrella Gymnastics', 'EST', '$35'],
    ['Houston Gymnastics Academy', 'HGA', '$20'],
    ['Oasis Gymnastics', 'OAS', '$20'],
    ['Rowland Ballard Atascocita', 'RBA', '$20'],
    ['Rowland Ballard Kingwood', 'RBK', '$15'],
    ['Scottsdale Gymnastics', 'SGT', '$30'],
    ['TIGAR Gymnastics', 'TIG', '$20'],
]

for i, row_data in enumerate(og_data):
    r = 4 + i
    ws2.cell(row=r, column=1, value=row_data[0])
    ws2.cell(row=r, column=2, value=row_data[1])
    ws2.cell(row=r, column=3, value=row_data[2])
    ws2.cell(row=r, column=4, value='No change')
    style_row(ws2, r, 4, no_change_fill)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal='left')

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 14

# ========== SHEET 3: SIBLING DISCOUNTS ==========
ws3 = wb.create_sheet('Sibling Discounts')

ws3.merge_cells('A1:D1')
ws3['A1'] = 'KNO SIBLING DISCOUNTS'
ws3['A1'].font = Font(bold=True, size=14, name='Arial', color='4A2C2A')
ws3['A1'].alignment = Alignment(horizontal='center')

r = 3
for i, h in enumerate(['Gym', 'Gym Code', 'KNO Sibling Discount', 'Status'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 4)

sib_data = [
    ['Capital Gymnastics Cedar Park', 'CCP', '$30 per sibling', 'Confirmed'],
    ['Capital Gymnastics Pflugerville', 'CPF', '$30 for 2nd child', 'Confirmed'],
    ['Capital Gymnastics Round Rock', 'CRR', '$30 per sibling', 'Confirmed'],
    ['Estrella Gymnastics', 'EST', 'Waiting on GM', 'Waiting'],
    ['Houston Gymnastics Academy', 'HGA', '$35 each for 2+ children', 'Confirmed'],
    ['Oasis Gymnastics', 'OAS', 'Waiting on GM', 'Waiting'],
    ['Rowland Ballard Atascocita', 'RBA', 'No sibling discounts', 'Confirmed'],
    ['Rowland Ballard Kingwood', 'RBK', 'Waiting on GM', 'Waiting'],
    ['Scottsdale Gymnastics', 'SGT', 'Waiting on GM', 'Waiting'],
    ['TIGAR Gymnastics', 'TIG', 'No sibling discounts', 'Confirmed'],
]

for i, row_data in enumerate(sib_data):
    r = 4 + i
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    fill = waiting_fill if row_data[3] == 'Waiting' else confirmed_fill
    style_row(ws3, r, 4, fill)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='left')

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 14

wb.save(r'C:\Users\Jayme\Downloads\Rise_Athletics_Pricing_Before_After.xlsx')
print('Saved to Downloads')
